# import matplotlib
import matplotlib.pyplot as plt
import pandas as pd
# import numpy as np
import re
import time
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

plt.rcParams['font.sans-serif'] = ['SimHei'] # for Chinese lable
plt.rcParams['axes.unicode_minus'] = False # for Chinese -

filename = 'chat_log.txt'
saints_excel_name = 'saints.xlsx'
data_str = '2020-\d-\d'
data_str2 = '2020-\d-\d\d'
time_str = '\d\d:\d\d'
plan_date_str = '2020\d\d\d\d'
saint_name_list = ['Anna', 'Carol', 'Figo', 'Grace', 'Jerry', 'Leon', 'Linda', 'Mandy', 'Siyuan', 'Sprindy']
saint_alias_list = ['Anna', 'Carol', 'WOW-LOL', 'GraceXu', 'blue sky', 'Leon', 'linda', 'Mandy', '周森森', 'Sprindy']

name_lists_by_time = [ [] for i in range(24)]
header = ['real_date', 'real_time', 'alias', 'plan_date', 'saint', 'real_time_int',
            saint_name_list[0], saint_name_list[1], saint_name_list[2], saint_name_list[3],
            saint_name_list[4], saint_name_list[5], saint_name_list[6], saint_name_list[7],
            saint_name_list[8], saint_name_list[9]]
saint_daily_record = header
study_data = pd.DataFrame(columns=header)

# study_data = pd.DataFrame(columns=['real_date', 'real_time', 'alias', 'plan_date', 'saint'],
#                            dtype=np.dtype(['datetime64', 'datetime64', 'str', 'datetime64', 'str']))

def search_and_save_data_to_excel():
    # flags for logic control
    is_real_date_found = False
    is_real_time_found = False
    is_plan_date_found = False
    is_the_saint_found = False

    with open(filename, 'r', encoding='utf8') as f:
        lines = f.readlines()
        line = f.readline()
        for line in lines:
            #value = [float(s) for s in line.split()]#4
            m = re.search(data_str, line)
            if m is not None:
                is_real_date_found = True
                # real_date
                n = re.search(data_str2, line)
                if n is not None:
                    saint_daily_record[0] = n.group()
                else:
                    saint_daily_record[0] = m.group()

            m = re.search(time_str, line)
            if m is not None:
                is_real_time_found = True
                # real_time
                saint_daily_record[1] = m.group()
                time_int = m.group().split(':')
                # print(time_int)
                saint_daily_record[5] = int(time_int[0]) + float(int(time_int[1])/60)
                # alias
                for i in range(len(saint_name_list)):
                    m = re.search(saint_alias_list[i], line, re.I)
                    if m is not None:
                        saint_daily_record[2] = m.group()
                        # print(saint_daily_record)


            m = re.search(plan_date_str, line)
            if m is not None:
                is_plan_date_found = True
                # plan_date
                saint_daily_record[3] = m.group()

            if (is_real_time_found and is_plan_date_found):
                for i in range(len(saint_name_list)):
                    m = re.search(saint_name_list[i], line, re.I)
                    if m is not None:
                        if (saint_daily_record[2] == saint_alias_list[i]):
                            # print(saint_alias_list[i])
                            is_the_saint_found = True
                            # saint
                            saint_daily_record[4] = m.group()
                            for j in range(len(saint_name_list)):
                                if (j == i):
                                    saint_daily_record[6+j] = saint_daily_record[5]
                                else:
                                    saint_daily_record[6+j] = ''

            # all found, clear for the next saint
            if (is_plan_date_found and is_the_saint_found):
                study_data.loc[len(study_data)] = saint_daily_record
                is_plan_date_found = False
                is_real_date_found = False
                is_the_saint_found = False

    # print(study_data)
    # print(study_data.dtypes)
    study_data['real_date'] = pd.to_datetime(study_data['real_date']).dt.date
    # study_data['real_time'] = pd.to_datetime(study_data['real_time']).dt.time
    # study_data['real_time'] = matplotlib.dates.date2num(study_data['real_time'])
    # print(study_data.dtypes)
    # save DataFrame to excel file
    study_data.to_excel(saints_excel_name)

def plot_with_data_from_excel():
    all_data = pd.read_excel(saints_excel_name)
    # print(reliable_data.iloc[:, 10].std())
    # plt.bar(reliable_data.index, reliable_data.AZ)
    # times = matplotlib.dates.date2num(all_data.real_time)
    # matplotlib.pyplot.plot_date(dates, values)
    plt.scatter(all_data.real_date, all_data.real_time_int)

    plt.show()

def plot_inside_excel():
    wb = load_workbook(filename=saints_excel_name)
    # Active WorkSheet
    ws = wb.active

    chat1 = ScatterChart()
    # style = MinMax(allow_none=True, min=1, max=48)
    # chat1.style = 7
    chat1.title = '2020 One Year Bible Study'
    chat1.x_axis.title = 'Date'
    chat1.y_axis.title = 'Time(o\'clock)'
    # enlarge the chart, default is too small
    # width = 15 # in cm, approx 5 rows
    # height = 7.5 # in cm, approx 14 rows
    chat1.height = chat1.height + 8
    chat1.width = chat1.width + 32

    xvalues = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
    for i in range(len(saint_name_list)):
        values = Reference(ws, min_col=8+i, min_row=2, max_row=ws.max_row)
        series = Series(values, xvalues, title=saint_name_list[i], title_from_data=False)
    #  {'triangle', 'dash', 'x', 'auto', 'diamond', 'circle', 'star',
    #   'picture', 'square', 'dot', 'plus'}
        series.marker = openpyxl.chart.marker.Marker('circle')
        series.graphicalProperties.line.noFill = True
        chat1.series.append(series)

    ws.add_chart(chat1, "A10")
    wb.save(filename=saints_excel_name)

# Main Entry
def main(args):
    search_and_save_data_to_excel()
    # plot_with_data_from_excel()
    plot_inside_excel()

if __name__ == '__main__':
  main(sys.argv)
